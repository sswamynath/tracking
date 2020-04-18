import openpyxl

import math

import os

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

wb = openpyxl.load_workbook('covid.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')

rowNo = 1
columnNo = 2
dateList = []

def display(loc, names,*args):

    for i in range(2, len(sheet['1'])):

        p = str(sheet.cell(row=1, column=i).value)
        dateList.append(p)
    
    if loc in set(dateList):
        columnNo = dateList.index(loc)+2

    for i in range(2, len(sheet['A'])):
        p = str(sheet.cell(row=i, column=1).value)
        if p == names:
            rowNo = i
    
    p = ""

    for i in range(int(len(args)/2)):  

        if args[2*i] != "" and args[2*i+1] != "":
            p = p+args[2*i]+","+args[2*i+1]+","

    p = p[:-1]

    r = p.upper()

    q = str(sheet.cell(row = rowNo, column = columnNo).value).upper()

    if q != "NONE" and p!="":
        p = q+","+p
        

    if p!="" and p!=q:    
        sheet.cell(row = rowNo, column = columnNo).value = p.upper()

    wb.save("covid.xlsx") 

    return q,r
            

    