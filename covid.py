import openpyxl

import math

import os

wb = openpyxl.load_workbook('covid.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')

def probVal(x, y):
    p = y
    q = 0
    for i in range(x-1):
        q = q+p/2
        p = p/2

    q = float("%.2f" % q)
    return q

def remarks(x):
    p = ""
    if x > 15:
        p = "Lockdown and Sanitization"
    elif x<=15 and x > 5:
        p = "Mandatory Sanitization"
    elif x<=5:
        p = "No Sanitization Required"
    return p

def times(x):
    p = x.split(", ")
    pl = []
    pn = []
    pt = []
    r = ""
    if len(p) != 0:
        for i in range(len(p)):
            if p[i] not in pl:
                pl.append(p[i])
                pn.append(1)
            elif p[i] in set(pl):
                pn[pl.index(p[i])] +=1

        pp = list(zip(pl, pn))
        pp = sorted(pp, key=lambda tup:(-tup[1], tup[0]))

        for i in range(len(pl)):
            q = str(pp[i][0])+" (x"+str(pp[i][1])+")"
            pt.append(q)
        for i in range(len(pt)):
            r = r+str(pt[i])+", "
        
        r = r[:-2]    
    return r



def display(pers, date, days):

    a = ''
    p = 0

    day = 2

    dateList = []

    for i in range(2, len(sheet['1'])):

        p = str(sheet.cell(row=1, column=i).value)
        dateList.append(p)
    
    if date in set(dateList):
        day = int(dateList.index(date)+2)

    period = int(days)
 

    iso = []
    isoProb = []
    isoLoc = []
    

    qua = []
    quaProb = []
    quaLoc = []

    obs = []
    obsProb = []
    obsLoc = []

    loc = []
    locProb = []
    locIso = []
    locQua = []
    locObs = []

    name = pers



    for i in range(2, len(sheet['A'])):

        if str(sheet.cell(row=i, column = 1).value) == name:
            for j in range(day-period, day):

                if str(sheet.cell(row=i, column=j).value) != "None":
                    q = str(sheet.cell(row=i, column=j).value)
                    qList = q.split(",")
                    # print(qList)
                    if len(qList)%2 != 0:
                        qList = qList[:-1]
                    # print(qList)
                    for k in range(len(qList)):
                        if k%2 == 0:
                            if qList[k] not in iso:
                                if qList[k] != pers:
                                    iso.append(qList[k])
                                    isoProb.append(1)
                            elif qList[k] in set(iso):
                                isoProb[iso.index(qList[k])] +=1
                        elif k%2 != 0:
                            if qList[k-1] in set(iso):
                                if isoProb[iso.index(qList[k-1])] == 1:
                                    isoLoc.append(qList[k])
                                    if qList[k] not in loc:
                                        loc.append(qList[k])
                                        locProb.append(5)
                                        locIso.append(1)
                                        locQua.append(0)
                                        locObs.append(0)
                                    elif qList[k] in set(loc):
                                        locProb[loc.index(qList[k])] +=5
                                        locIso[loc.index(qList[k])] +=1
                                    
                                else:
                                    isoLoc[iso.index(qList[k-1])] = isoLoc[iso.index(qList[k-1])]+str(', ')+qList[k]
                                    if qList[k] not in loc:
                                        loc.append(qList[k])
                                        locProb.append(5)
                                        locIso.append(1)
                                        locQua.append(0)
                                        locObs.append(0)
                                    elif qList[k] in set(loc):
                                        locProb[loc.index(qList[k])] +=5
                                        locIso[loc.index(qList[k])] +=1

    if len(iso) != 0:   
        for i in range(len(iso)):
            name=iso[i]  
            for i in range(2, len(sheet['A'])):

                if str(sheet.cell(row=i, column = 1).value) == name:
                    for j in range(day-period+2, day+2):
                        if str(sheet.cell(row=i, column=j).value) != "None":
                            q = str(sheet.cell(row=i, column=j).value)
                            qList = q.split(",")
                            if len(qList)%2 != 0:
                                qList = qList[:-1]
                            for k in range(len(qList)):
                                if k%2 == 0:
                                    if qList[k] not in iso:
                                        if qList[k] not in qua:
                                            if qList[k] != pers:
                                                qua.append(qList[k])
                                                quaProb.append(1)
                                        elif qList[k] in set(qua):
                                            quaProb[qua.index(qList[k])] +=1
                                elif k%2 != 0:
                                    if qList[k-1] in set(qua):
                                        if quaProb[qua.index(qList[k-1])] == 1:
                                            quaLoc.append(qList[k])
                                            if qList[k] not in loc:
                                                loc.append(qList[k])
                                                locProb.append(3)
                                                locIso.append(0)
                                                locQua.append(1)
                                                locObs.append(0)
                                            elif qList[k] in set(loc):
                                                locProb[loc.index(qList[k])] +=3
                                                locQua[loc.index(qList[k])] +=1
                                        else:
                                            # print(qua.index(qList[k-1]))
                                            # print(len(quaLoc))
                                            quaLoc[qua.index(qList[k-1])] = quaLoc[qua.index(qList[k-1])]+str(', ')+qList[k]
                                            if qList[k] not in loc:
                                                loc.append(qList[k])
                                                locProb.append(3)
                                                locIso.append(0)
                                                locQua.append(1)
                                                locObs.append(0)
                                            elif qList[k] in set(loc):
                                                locProb[loc.index(qList[k])] +=3
                                                locQua[loc.index(qList[k])] +=1
                else:
                    pass
                            
    if len(qua) != 0:
        for i in range(len(qua)):
            name=qua[i]  
            for i in range(2, len(sheet['A'])):

                if str(sheet.cell(row=i, column = 1).value) == name:
                    for j in range(day-period+2, day+2):
                        if str(sheet.cell(row=i, column=j).value) != "None":
                            q = str(sheet.cell(row=i, column=j).value)
                            qList = q.split(",")
                            if len(qList)%2 != 0:
                                qList = qList[:-1]
                            for k in range(len(qList)):
                                if k%2 == 0:
                                    if qList[k] not in iso:
                                        if qList[k] not in qua:
                                            if qList[k] not in obs:
                                                if qList[k] != pers:
                                                    obs.append(qList[k])
                                                    obsProb.append(1)
                                            elif qList[k] in set(obs):
                                                obsProb[obs.index(qList[k])] +=1
                                elif k%2 != 0:
                                    if qList[k-1] in set(obs):
                                        if obsProb[obs.index(qList[k-1])] == 1:
                                            obsLoc.append(qList[k])
                                            if qList[k] not in loc:
                                                loc.append(qList[k])
                                                locProb.append(1)
                                                locIso.append(0)
                                                locQua.append(0)
                                                locObs.append(1)
                                            elif qList[k] in set(loc):
                                                locProb[loc.index(qList[k])] +=1
                                                locObs[loc.index(qList[k])] +=1
                                        else:
                                            obsLoc[obs.index(qList[k-1])] = obsLoc[obs.index(qList[k-1])]+str(', ')+qList[k]
                                            if qList[k] not in loc:
                                                loc.append(qList[k])
                                                locProb.append(1)
                                                locIso.append(0)
                                                locQua.append(0)
                                                locObs.append(1)
                                            elif qList[k] in set(loc):
                                                locProb[loc.index(qList[k])] +=1
                                                locObs[loc.index(qList[k])] +=1
                else:
                    pass
                                    

    # print(iso, isoLoc, isoProb)
    # print(len(iso), len(isoLoc), len(isoProb))
    # print(qua)
    # print(obs)

    isoTup = list(zip(iso, isoProb, isoLoc))
    # print(isoTup)
    isoTup = sorted(isoTup, key=lambda tup:(-tup[1], tup[0]))
    # print(isoTup)

    quaTup = list(zip(qua, quaProb, quaLoc))
    quaTup = sorted(quaTup, key=lambda tup:(-tup[1], tup[0]))

    obsTup = list(zip(obs, obsProb, obsLoc))
    obsTup = sorted(obsTup, key=lambda tup:(-tup[1], tup[0]))

    locTup = list(zip(loc, locProb, locIso, locQua, locObs))
    locTup = sorted(locTup, key=lambda tup:(-tup[1], tup[0]))
            

    text = ''
    for i in range(len(isoLoc)):
        text = text+str(i+1)+'\t'+isoTup[i][0].title()+'\t'+'\t'+'\t'+'\t'+str(isoTup[i][1])+'\t'+str(60+probVal(isoTup[i][1],40))+str('%')+'\t'+str(times(str(isoTup[i][2]))).title()+'\n'

    # print(text)

    textq = ''
    for i in range(len(quaLoc)):
        textq = textq+str(i+1)+'\t'+quaTup[i][0].title()+'\t'+'\t'+'\t'+'\t'+str(quaTup[i][1])+'\t'+str(20+probVal(quaTup[i][1],40))+str('%')+'\t'+str(times(str(quaTup[i][2]))).title()+'\n'

    # print(textq)

    texto = ''
    for i in range(len(obsLoc)):
        texto = texto+str(i+1)+'\t'+obsTup[i][0].title()+'\t'+'\t'+'\t'+'\t'+str(obsTup[i][1])+'\t'+str(probVal(obsTup[i][1],20))+str('%')+'\t'+str(times(str(obsTup[i][2]))).title()+'\n'

    # print(texto)

    textl = ''
    for i in range(len(loc)):
        textl = textl+str(i+1)+'\t'+locTup[i][0].title()+'\t'+'\t'+'\t'+'\t'+str(locTup[i][2])+'\t'+str(locTup[i][3])+'\t'+str(locTup[i][4])+'\t'+str(locTup[i][1])+'\t'+str(remarks(locTup[i][1]))+'\n'

    # print(texto)

    f= open("covid.txt","w+")
    f.write("SPREAD OF CORONA VIRUS"+"\n"+"\n")
    f.write("Total number of Pers for Isolation : "+str(len(iso))+"\n")
    f.write(text)
    f.write("\n"+"Total number of Pers for Quarantine : "+str(len(qua))+"\n")
    f.write(textq)
    f.write("\n"+"Total number of Pers for Observation : "+str(len(obs))+"\n")
    f.write(texto)
    f.write("\n"+"Rank of Place : "+"\n")
    f.write(textl)
    f.close() 

    if len(date.split())>3:
        os.startfile("covid.txt")


    return text, len(iso), textq, len(qua), texto, len(obs), textl, len(loc)