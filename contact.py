from tkinter import *

import tkinter.font as tkFont

from tkinter import ttk

import covid

import dataEntry

import math

import os

import openpyxl

import tkentrycomplete

window = Tk()

window.title("201 Engineer Regiment Covid-19 Contact Tracking Software")

window.wm_iconbitmap('unit.ico')

default_font = tkFont.nametofont("TkDefaultFont")

default_font.configure(size=15)

window.option_add("*Font", default_font)

# window.attributes('-fullscreen', True)

window.state('zoomed')

tab_control = ttk.Notebook(window)

tab1 = ttk.Frame(tab_control)
 
tab2 = ttk.Frame(tab_control)

tab_control.add(tab1, text = '    QURANTINE ANALYSIS    ')
 
tab_control.add(tab2, text=  '    CONTACT REGISTRATION    ')

tab_control.pack(expand=1, fill='both')


class covidClass:
    def argsToCovid(self):
        self.pers = str(guiCovid.names.get())
        self.date = str(guiCovid.date.get())
        self.days = int(guiCovid.days.get())
        # print(self.pers, self.date, self.days)
        a = covid.display(self.pers, self.date, self.days)

        guiCovid.bBoundsi.delete('1.0', END)
        guiCovid.bBoundsq.delete('1.0', END)
        guiCovid.bBoundso.delete('1.0', END)
        guiCovid.bBoundsl.delete('1.0', END)
        

        b= "Personnel Recommended for Isolation : "+str(a[1])+"\n"
        c= "Personnel Recommended for Quarantine : "+str(a[3])+"\n"
        d= "Personnel Recommended for Observation : "+str(a[5])+"\n"
        # e= "S.No "+"\t"+"Name"+"\t"+"\t"+"\t"+"\t"+"Times"+"\t"+"Probability"+"\t"+"Location"+"\n"

        guiCovid.bBoundsi.insert(INSERT, b)
        # guiCovid.bBoundsi.insert(INSERT, e)
        guiCovid.bBoundsi.insert(INSERT, a[0])

        guiCovid.bBoundsq.insert(INSERT, c)
        guiCovid.bBoundsq.insert(INSERT, a[2])

        guiCovid.bBoundso.insert(INSERT, d)
        guiCovid.bBoundso.insert(INSERT, a[4])

        guiCovid.bBoundsl.insert(INSERT, a[6])

    def argsToCovidClear(self):

        guiCovid.names.set("")
        guiCovid.date.set("")     
        guiCovid.bBoundsi.delete('1.0', END)
        guiCovid.bBoundsq.delete('1.0', END)
        guiCovid.bBoundso.delete('1.0', END)
        guiCovid.bBoundsl.delete('1.0', END)


class dataClass:
    def argsToDataEntry(self):

        self.loc = str(guiData.loc.get())
        self.names = str(guiData.names.get())
        self.loc1 = str(guiData.loc1.get())
        self.names1 = str(guiData.names1.get())
        self.loc2 = str(guiData.loc2.get())
        self.names2 = str(guiData.names2.get())
        self.loc3 = str(guiData.loc3.get())
        self.names3 = str(guiData.names3.get())
        self.loc4 = str(guiData.loc4.get())
        self.names4 = str(guiData.names4.get())
        self.loc5 = str(guiData.loc5.get())
        self.names5 = str(guiData.names5.get())
        self.loc6 = str(guiData.loc6.get())
        self.names6 = str(guiData.names6.get())
        self.loc7 = str(guiData.loc7.get())
        self.names7 = str(guiData.names7.get())
        self.loc8 = str(guiData.loc8.get())
        self.names8 = str(guiData.names8.get())
        self.loc9 = str(guiData.loc9.get())
        self.names9 = str(guiData.names9.get())
        self.loc10 = str(guiData.loc10.get())
        self.names10 = str(guiData.names10.get())
        self.loc11 = str(guiData.loc11.get())
        self.names11 = str(guiData.names11.get())
        self.loc12 = str(guiData.loc12.get())
        self.names12 = str(guiData.names12.get())
        self.loc13 = str(guiData.loc13.get())
        self.names13 = str(guiData.names13.get())
        self.loc14 = str(guiData.loc14.get())
        self.names14 = str(guiData.names14.get())
        self.loc15 = str(guiData.loc15.get())
        self.names15 = str(guiData.names15.get())
        self.loc16 = str(guiData.loc16.get())
        self.names16 = str(guiData.names16.get())      
        

        a = dataEntry.display(self.loc, self.names,self.names1,self.loc1, self.names2, self.loc2, self.names3,self.loc3, self.names4,self.loc4, self.names5,self.loc5, self.names6,self.loc6, self.names7,self.loc7, self.names8, self.loc8, self.names9, self.loc9, self.names10, self.loc10, self.names11, self.loc11, self.names12, self.loc12, self.names13, self.loc13, self.names14, self.loc14, self.names15, self.loc15, self.names16, self.loc16)
        b = "Previous Entry : "+str(a[0])+"\n"+"Updated Entry : "+str(a[1])
        guiData.bBounds.delete('1.0', END)
        guiData.bBounds.insert(INSERT, b)

    def argsToClearData(self):
        guiData.loc.set('')
        guiData.names.set('')
        guiData.names1.set('')
        guiData.loc1.set('')
        guiData.names2.set('')
        guiData.loc2.set('')
        guiData.names3.set('')
        guiData.loc3.set('')
        guiData.names4.set('')
        guiData.loc4.set('')
        guiData.names5.set('')
        guiData.loc5.set('')
        guiData.names6.set('')
        guiData.loc6.set('')
        guiData.names7.set('')
        guiData.loc7.set('')
        guiData.names8.set('')
        guiData.loc8.set('')
        guiData.names9.set('')
        guiData.loc9.set('')
        guiData.names10.set('')
        guiData.loc10.set('')
        guiData.names11.set('')
        guiData.loc11.set('')
        guiData.names12.set('')
        guiData.loc12.set('')
        guiData.names13.set('')
        guiData.loc13.set('')
        guiData.names14.set('')
        guiData.loc14.set('')
        guiData.names15.set('')
        guiData.loc15.set('')
        guiData.names16.set('')
        guiData.loc16.set('')


        guiData.bBounds.delete('1.0', END)



class guiCovid:

    titleFrame = Frame(tab1,bg = "black")
    titleFrame.grid(row = 0, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    title = Label(titleFrame, text = "ERUNOOTRU ONNE"+"\n"+"COVID-19 CONTACT TRACKING SOFTWARE", fg = 'yellow', bg = "black", width = 48)
    title.config(font=("Courier", 30, "bold"))
    title.grid(row = 0, column = 1, sticky = N, columnspan = 4)

    img = PhotoImage(file='unit.png')

    title1 = Label(titleFrame, image = img, bg = "black")
    title1.grid(row = 0, column = 0, sticky = W)

    title2 = Label(titleFrame, image = img, bg = "black")
    title2.grid(row = 0, column = 5, sticky = E)
    
    parameters = LabelFrame(tab1, text = "Parameters")
    parameters.grid(row = 1, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    names = StringVar(parameters)
    date = StringVar(parameters)
    days = StringVar(parameters)

    nameList = []
    dateList = []
    daysList = []

    wb = openpyxl.load_workbook('covid.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    for i in range(2, len(sheet['A'])):

        p = str(sheet.cell(row=i, column=1).value)
        nameList.append(p)   

    for i in range(2, len(sheet['1'])):

        p = str(sheet.cell(row=1, column=i).value)
        dateList.append(p) 

    for i in range(1, 31):

        daysList.append(i) 

    lPers = Label(parameters, text = "Name : ",  anchor = W)
    lPers.grid(row = 0, column = 0, sticky = W, padx = 10)

    ePers = tkentrycomplete.AutocompleteCombobox(parameters,width = 25, textvariable=names)
    ePers.set_completion_list(nameList)
    ePers.grid(row=0, column =1,sticky = W, padx = 10)

    lDate = Label(parameters, text = "Date : ",  anchor = W)
    lDate.grid(row = 0, column = 2, sticky = W, padx = 10)

    eDate = tkentrycomplete.AutocompleteCombobox(parameters,width = 15, textvariable=date)
    eDate.set_completion_list(dateList)
    eDate.grid(row=0, column =3,sticky = W, padx = 10)

    lDays = Label(parameters, text = "No of Days : ",  anchor = W)
    lDays.grid(row = 0, column = 4, sticky = W, padx = 10)

    eDays = ttk.Combobox(parameters,textvariable=days,width = 5, value=daysList)
    eDays.grid(row=0, column =5,sticky = W, padx = 10)

    covidObj = covidClass()

    bPreview = Button(parameters, text="Update",width = 12,bg = 'green', command = covidObj.argsToCovid)
    bPreview.grid(row=0, column = 6, padx = 40)

    bPreviewc = Button(parameters, text="Clear",width = 12,bg = 'red', command = covidObj.argsToCovidClear)
    bPreviewc.grid(row=0, column = 7, padx = 0)

    isolation = LabelFrame(tab1, text = "Isolation (Primary Contact)")
    isolation.grid(row = 3, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    bBoundsi = Text(isolation, wrap=WORD, width=120, height= 4)         
    bBoundsi.grid(row = 6, column = 0, columnspan = 4, padx = 10, pady = 5)


    quarantine = LabelFrame(tab1, text = "Quarantine (Secondary Contact)")
    quarantine.grid(row = 4, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    bBoundsq = Text(quarantine, wrap=WORD, width=120, height= 3)      
    bBoundsq.grid(row = 6, column = 0, columnspan = 4, padx = 10, pady = 5)

    observation = LabelFrame(tab1, text = "Under Observation (Teritary Contact)")
    observation.grid(row = 5, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    bBoundso = Text(observation, wrap=WORD, width=120, height= 3)       
    bBoundso.grid(row = 6, column = 0, columnspan = 4, padx = 10, pady = 5)

    location = LabelFrame(tab1, text = "Locations Affected")
    location.grid(row = 6, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    bBoundsl = Text(location, wrap=WORD, width=120, height= 2)       
    bBoundsl.grid(row = 6, column = 0, columnspan = 4, padx = 10, pady = 5)

class guiData:

    titleFrame = Frame(tab2,bg = "black")
    titleFrame.grid(row = 0, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    title = Label(titleFrame, text = "ERUNOOTRU ONNE"+"\n"+"COVID-19 CONTACT TRACKING SOFTWARE", fg = 'yellow', bg = "black", width = 48)
    title.config(font=("Courier", 30, "bold"))
    title.grid(row = 0, column = 1, sticky = N, columnspan = 4)

    img = PhotoImage(file='unit.png')

    title1 = Label(titleFrame, image = img, bg = "black")
    title1.grid(row = 0, column = 0, sticky = W)

    title2 = Label(titleFrame, image = img, bg = "black")
    title2.grid(row = 0, column = 5, sticky = E)

    pers = LabelFrame(tab2, text = "Select Person For Data Entry")
    pers.grid(row = 1, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    contact = LabelFrame(tab2, text = "Contact History")
    contact.grid(row = 2, column = 0, sticky = W+E, columnspan=  2, padx = 5)

    contact1 = LabelFrame(tab2, text = "Contact History")
    contact1.grid(row = 2, column = 2, sticky = W+E, columnspan=  2, padx = 5)

    result = Frame(tab2)
    result.grid(row = 4, column = 0, sticky = W+E, columnspan=  4, padx = 5)

    lNames = Label(pers, text = "Enter Name", width = 30, anchor = W)
    lNames.grid(row = 0, column = 0, sticky = W, padx = 10)
    
    lLoc = Label(pers, text = "Enter Date", width = 30, anchor = W)
    lLoc.grid(row = 0, column = 2, sticky = W, padx = 10)    

    lNames1 = Label(contact, text = "Contact Personnel Name",width = 28, anchor = W)
    lNames1.grid(row = 0, column = 0, sticky = W, padx = 10)

    lLoc1 = Label(contact, text = "Contact Location", width = 28,anchor = W)
    lLoc1.grid(row = 0, column = 1, sticky = W, padx = 10)

    lNames2 = Label(contact1, text = "Contact Personnel Name",width = 28, anchor = W)
    lNames2.grid(row = 0, column = 0, sticky = W, padx = 10)

    lLoc2 = Label(contact1, text = "Contact Location", width = 28,anchor = W)
    lLoc2.grid(row = 0, column = 1, sticky = W, padx = 10)

    wb = openpyxl.load_workbook('covid.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    sheet1 = wb.get_sheet_by_name('Sheet2')

    nameList = []
    dateList = []
    locList = []

    names = StringVar(contact)
    names1 = StringVar(contact)
    names2 = StringVar(contact)
    names3 = StringVar(contact)
    names4 = StringVar(contact)
    names5 = StringVar(contact)
    names6 = StringVar(contact)
    names7 = StringVar(contact)
    names8 = StringVar(contact)
    names9 = StringVar(contact)
    names10 = StringVar(contact)
    names11 = StringVar(contact)
    names12 = StringVar(contact)
    names13 = StringVar(contact)
    names14 = StringVar(contact)
    names15 = StringVar(contact)
    names16 = StringVar(contact)

    loc = StringVar(contact)
    loc1 = StringVar(contact)
    loc2 = StringVar(contact)
    loc3 = StringVar(contact)
    loc4 = StringVar(contact)
    loc5 = StringVar(contact)
    loc6 = StringVar(contact)
    loc7 = StringVar(contact)
    loc8 = StringVar(contact)
    loc9 = StringVar(contact)
    loc10 = StringVar(contact)
    loc11 = StringVar(contact)
    loc12 = StringVar(contact)
    loc13 = StringVar(contact)
    loc14 = StringVar(contact)
    loc15 = StringVar(contact)
    loc16 = StringVar(contact)
    


    for i in range(2, len(sheet['A'])):

        p = str(sheet.cell(row=i, column=1).value)
        nameList.append(p)
    
    for i in range(2, len(sheet['1'])):

        p = str(sheet.cell(row=1, column=i).value)
        dateList.append(p)

    for i in range(2, len(sheet1['A'])):

        p = str(sheet1.cell(row=i, column=1).value)
        locList.append(p)


    nameOption = tkentrycomplete.AutocompleteCombobox(pers, textvariable=names)
    nameOption.set_completion_list(nameList)
    nameOption.grid(row=0, column =1,sticky = N, padx = 10)

    locOption = tkentrycomplete.AutocompleteCombobox(pers, textvariable=loc)
    locOption.set_completion_list(dateList)
    locOption.grid(row=0, column =3,sticky = N, padx = 10)    

    nameOption1 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names1)
    nameOption1.set_completion_list(nameList)
    nameOption1.grid(row=1, column =0,sticky = W, padx = 10, pady = 4)
    
    locOption1 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc1)
    locOption1.set_completion_list(locList)
    locOption1.grid(row=1, column =1,sticky = W, padx = 10, pady = 4)

    nameOption2 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names2)
    nameOption2.set_completion_list(nameList)
    nameOption2.grid(row=2, column =0,sticky = W, padx = 10, pady = 4)

    locOption2 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc2)
    locOption2.set_completion_list(locList)
    locOption2.grid(row=2, column =1,sticky = W, padx = 10, pady = 4)

    nameOption3 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names3)
    nameOption3.set_completion_list(nameList)
    nameOption3.grid(row=3, column =0,sticky = W, padx = 10, pady = 4)

    locOption3 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc3)
    locOption3.set_completion_list(locList)
    locOption3.grid(row=3, column =1,sticky = W, padx = 10, pady = 4)

    nameOption4 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names4)
    nameOption4.set_completion_list(nameList)
    nameOption4.grid(row=4, column =0,sticky = W, padx = 10, pady = 4)

    locOption4 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc4)
    locOption4.set_completion_list(locList)
    locOption4.grid(row=4, column =1,sticky = W, padx = 10, pady = 4)

    nameOption5 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names5)
    nameOption5.set_completion_list(nameList)
    nameOption5.grid(row=5, column =0,sticky = W, padx = 10, pady = 4)

    locOption5 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc5)
    locOption5.set_completion_list(locList)
    locOption5.grid(row=5, column =1,sticky = W, padx = 10, pady = 4)

    nameOption6 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names6)
    nameOption6.set_completion_list(nameList)
    nameOption6.grid(row=6, column =0,sticky = W, padx = 10, pady = 4)

    locOption6 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc6)
    locOption6.set_completion_list(locList)
    locOption6.grid(row=6, column =1,sticky = W, padx = 10, pady = 4)

    nameOption7 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names7)
    nameOption7.set_completion_list(nameList)
    nameOption7.grid(row=7, column =0,sticky = W, padx = 10, pady = 4)

    locOption7 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc7)
    locOption7.set_completion_list(locList)
    locOption7.grid(row=7, column =1,sticky = W, padx = 10, pady = 4)

    nameOption8 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=names8)
    nameOption8.set_completion_list(nameList)
    nameOption8.grid(row=8, column =0,sticky = W, padx = 10, pady = 4)

    locOption8 = tkentrycomplete.AutocompleteCombobox(contact, textvariable=loc8)
    locOption8.set_completion_list(locList)
    locOption8.grid(row=8, column =1,sticky = W, padx = 10, pady = 4)

    nameOption9 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names9)
    nameOption9.set_completion_list(nameList)
    nameOption9.grid(row=1, column =0,sticky = W, padx = 10, pady = 4)

    locOption9 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc9)
    locOption9.set_completion_list(locList)
    locOption9.grid(row=1, column =1,sticky = W, padx = 10, pady = 4)

    nameOption10 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names10)
    nameOption10.set_completion_list(nameList)
    nameOption10.grid(row=2, column =0,sticky = W, padx = 10, pady = 4)

    locOption10 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc10)
    locOption10.set_completion_list(locList)
    locOption10.grid(row=2, column =1,sticky = W, padx = 10, pady = 4)

    nameOption11 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names11)
    nameOption11.set_completion_list(nameList)
    nameOption11.grid(row=3, column =0,sticky = W, padx = 10, pady = 4)

    locOption11 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc11)
    locOption11.set_completion_list(locList)
    locOption11.grid(row=3, column =1,sticky = W, padx = 10, pady = 4)

    nameOption12 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names12)
    nameOption12.set_completion_list(nameList)
    nameOption12.grid(row=4, column =0,sticky = W, padx = 10, pady = 4)

    locOption12 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc12)
    locOption12.set_completion_list(locList)
    locOption12.grid(row=4, column =1,sticky = W, padx = 10, pady = 4)

    nameOption13 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names13)
    nameOption13.set_completion_list(nameList)
    nameOption13.grid(row=5, column =0,sticky = W, padx = 10, pady = 4)

    locOption13 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc13)
    locOption13.set_completion_list(locList)
    locOption13.grid(row=5, column =1,sticky = W, padx = 10, pady = 4)

    nameOption14 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names14)
    nameOption14.set_completion_list(nameList)
    nameOption14.grid(row=6, column =0,sticky = W, padx = 10, pady = 4)

    locOption14 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc14)
    locOption14.set_completion_list(locList)
    locOption14.grid(row=6, column =1,sticky = W, padx = 10, pady = 4)

    nameOption15 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names15)
    nameOption15.set_completion_list(nameList)
    nameOption15.grid(row=7, column =0,sticky = W, padx = 10, pady = 4)

    locOption15 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc15)
    locOption15.set_completion_list(locList)
    locOption15.grid(row=7, column =1,sticky = W, padx = 10, pady = 4)

    nameOption16 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=names16)
    nameOption16.set_completion_list(nameList)
    nameOption16.grid(row=8, column =0,sticky = W, padx = 10, pady = 4)

    locOption16 = tkentrycomplete.AutocompleteCombobox(contact1, textvariable=loc16)
    locOption16.set_completion_list(locList)
    locOption16.grid(row=8, column =1,sticky = W, padx = 10, pady = 4)  

    dataObj = dataClass()

    bPreview = Button(tab2, text="Update",width = 20, bg = 'green',command = dataObj.argsToDataEntry)
    bPreview.grid(row=3, column = 1, padx = 80)

    bPreviewc = Button(tab2, text="Clear",width = 20, bg = 'red',command = dataObj.argsToClearData)
    bPreviewc.grid(row=3, column = 3, padx = 80)

    bBounds = Text(result, wrap=WORD, width=120, height= 3)         
    bBounds.grid(row = 1, column = 0, columnspan = 4, padx = 10, pady = 5)


window.mainloop()
